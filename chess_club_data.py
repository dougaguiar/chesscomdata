import requests
import pandas as pd
from collections import Counter, defaultdict
from datetime import datetime
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tempfile
import os

# --- CONFIGURATION ---
CLUB_NAME = "torneios-cursos-gm-krikor"
HEADERS = {'User-Agent': 'your_username@example.com'}
BASE_URL = "https://api.chess.com/pub"
YEARS = [2024, 2024]
GAME_CLASSES = ['blitz', 'rapid']
FIRST_MONTH, LAST_MONTH = 1, 12
MEMBER_LIMIT = None

# --- DERIVED CONFIGURATION ---
FIRST_YEAR, LAST_YEAR = min(YEARS), max(YEARS)

# --- HELPER FUNCTIONS ---

def get_current_ratings(username):
    url = f"{BASE_URL}/player/{username}/stats"
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        data = response.json()
        return {tc: data.get(f"chess_{tc}", {}).get('last', {}).get('rating') for tc in GAME_CLASSES}
    except requests.exceptions.RequestException as e:
        print(f"[WARNING] Could not fetch current ratings for {username}: {e}")
        return {tc: None for tc in GAME_CLASSES}

def process_member(username):
    game_counts = defaultdict(int)
    white_openings = Counter()
    black_defenses = Counter()
    ratings_over_time = defaultdict(list)
    
    for year in range(FIRST_YEAR, LAST_YEAR + 1):
        for month in range(1, 13):
            if (year == FIRST_YEAR and month < FIRST_MONTH) or \
               (year == LAST_YEAR and month > LAST_MONTH):
                continue

            url = f"{BASE_URL}/player/{username}/games/{year}/{month:02d}"
            try:
                response = requests.get(url, headers=HEADERS)
                response.raise_for_status()
                games = response.json().get('games', [])

                for game in games:
                    time_class = game.get('time_class')
                    if time_class not in GAME_CLASSES:
                        continue
                    
                    game_counts[time_class] += 1
                    
                    pgn = game.get('pgn', '')
                    eco_name = 'Unknown Opening'
                    if pgn:
                        eco_url = next((line.split('"')[1] for line in pgn.split('\n') if line.startswith('[ECOUrl ')), None)
                        eco_name = eco_url.replace("https://www.chess.com/openings/", "").replace("-", " ").strip() if eco_url else next((line.split('"')[1] for line in pgn.split('\n') if line.startswith('[ECO ')), 'Unknown Opening')
                    
                    if eco_name != 'Unknown Opening':
                        player_color = 'white' if game['white']['username'].lower() == username.lower() else 'black'
                        (white_openings if player_color == 'white' else black_defenses)[eco_name] += 1

                    end_time = game.get('end_time')
                    player_rating = game.get(player_color, {}).get('rating')
                    if end_time and player_rating:
                        ratings_over_time[time_class].append((end_time, player_rating))

            except requests.exceptions.RequestException:
                continue

    total_games_analyzed = sum(game_counts.values())
    preferred_opening, opening_count = white_openings.most_common(1)[0] if white_openings else ("N/A", 0)
    preferred_defense, defense_count = black_defenses.most_common(1)[0] if black_defenses else ("N/A", 0)

    member_data = {
        'username': username, 'total_games': total_games_analyzed, 'blitz_games': game_counts['blitz'],
        'rapid_games': game_counts['rapid'], 'white_openings': white_openings, 'black_defenses': black_defenses,
        'preferred_white_opening': preferred_opening, 'white_opening_count': opening_count,
        'preferred_black_defense': preferred_defense, 'black_defense_count': defense_count,
        **get_current_ratings(username)
    }

    for tc in GAME_CLASSES:
        initial_rating, final_rating, increment = None, None, 0
        if ratings_over_time[tc]:
            sorted_ratings = sorted(ratings_over_time[tc])
            initial_rating, final_rating = sorted_ratings[0][1], sorted_ratings[-1][1]
            increment = final_rating - initial_rating
        member_data[f'initial_{tc}_rating'] = initial_rating
        member_data[f'final_{tc}_rating'] = final_rating
        member_data[f'increment_{tc}'] = increment
    
    return member_data

def auto_adjust_sheet_columns(worksheet, df):
    for col_idx, column in enumerate(df.columns, 1):
        max_len = max(len(str(column)), df[column].astype(str).map(len).max()) + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len

def generate_excel_report(member_stats, output_file):
    if not member_stats:
        return

    members_df = pd.DataFrame(member_stats)
    for tc in GAME_CLASSES:
        for col_suffix in ['_rating', 'increment_']:
            col_name = f'initial_{tc}{col_suffix}' if 'initial' in col_suffix else f'{col_suffix}{tc}'
            if col_name not in members_df: members_df[col_name] = 0
        if tc not in members_df: members_df[tc] = None

    # ✨ YOUR NEW METRICS ARE INCLUDED HERE
    summary_data = {
        'Club Name': CLUB_NAME, 'Analysis Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Analysis Period': f"{datetime(FIRST_YEAR, FIRST_MONTH, 1).strftime('%b %Y')} to {datetime(LAST_YEAR, LAST_MONTH, 1).strftime('%b %Y')}",
        'Total Members Processed': len(members_df), 'Total Games Analyzed': members_df['total_games'].sum(),
        'Total Rapid Games': members_df['rapid_games'].sum(), 'Total Blitz Games': members_df['blitz_games'].sum(),
        'Average Blitz Rating (End of Period)': f"{members_df['final_blitz_rating'].mean():.2f}" if 'final_blitz_rating' in members_df and not members_df['final_blitz_rating'].dropna().empty else "N/A",
        'Average Rapid Rating (End of Period)': f"{members_df['final_rapid_rating'].mean():.2f}" if 'final_rapid_rating' in members_df and not members_df['final_rapid_rating'].dropna().empty else "N/A",
        'Average Blitz Rating (Current)': f"{members_df['blitz'].mean():.2f}" if 'blitz' in members_df and not members_df['blitz'].dropna().empty else "N/A",
        'Average Rapid Rating (Current)': f"{members_df['rapid'].mean():.2f}" if 'rapid' in members_df and not members_df['rapid'].dropna().empty else "N/A",
        'Total Blitz Rating Change': f"{members_df['increment_blitz'].sum():+.0f}",
        'Total Rapid Rating Change': f"{members_df['increment_rapid'].sum():+.0f}",
    }
    summary_df = pd.DataFrame(list(summary_data.items()), columns=['Statistic', 'Value'])

    top10_active_df = members_df.sort_values(by='total_games', ascending=False).head(10)[['username', 'total_games', 'blitz_games', 'rapid_games']]
    
    total_white_openings, total_black_defenses = Counter(), Counter()
    for data in member_stats:
        total_white_openings.update(data['white_openings'])
        total_black_defenses.update(data['black_defenses'])
    
    openings_df = pd.DataFrame(total_white_openings.most_common(10), columns=['Opening', 'Frequency'])
    defenses_df = pd.DataFrame(total_black_defenses.most_common(10), columns=['Defense', 'Frequency'])

    fig, ax = plt.subplots(figsize=(10, 6))
    ratings_to_plot = [members_df[tc].dropna() for tc in GAME_CLASSES if tc in members_df and not members_df[tc].dropna().empty]
    if ratings_to_plot:
        ax.hist(ratings_to_plot, bins=range(400, 3000, 100), label=[tc.capitalize() for tc in GAME_CLASSES if tc in members_df and not members_df[tc].dropna().empty], alpha=0.7, edgecolor='white')
        ax.set_title('Club Rating Distribution (Current)', fontsize=16)
        ax.set_xlabel('Rating', fontsize=12); ax.set_ylabel('Number of Players', fontsize=12)
        ax.legend(); ax.grid(axis='y', linestyle='--', alpha=0.7)
    else:
        ax.text(0.5, 0.5, 'No rating data available.', ha='center', va='center')

    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
        plt.savefig(tmpfile.name, bbox_inches='tight'); plot_path = tmpfile.name
    plt.close(fig)

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            raw_data_columns = [
                'username', 'total_games', 'blitz_games', 'rapid_games', 'initial_blitz_rating', 'final_blitz_rating', 'increment_blitz',
                'initial_rapid_rating', 'final_rapid_rating', 'increment_rapid', 'blitz', 'rapid', 'preferred_white_opening', 
                'white_opening_count', 'preferred_black_defense', 'black_defense_count'
            ]
            members_display_df = members_df[[col for col in raw_data_columns if col in members_df.columns]]

            for df, sheet_name in [(summary_df, 'Summary'), (members_display_df, 'Members Raw Data'), 
                                   (top10_active_df, 'Top 10 Active'), (openings_df, 'Popular Openings'), 
                                   (defenses_df, 'Popular Defenses')]:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                auto_adjust_sheet_columns(writer.sheets[sheet_name], df)
            
            if os.path.exists(plot_path):
                img = XLImage(plot_path); img.anchor = 'D4'
                writer.sheets['Summary'].add_image(img)
    except Exception as e:
        print(f"\n[ERROR] Could not write to Excel file '{output_file}': {e}")
    finally:
        if os.path.exists(plot_path): os.remove(plot_path)

# --- SCRIPT EXECUTION ---
start_time = time.time()
try:
    club_url = f"{BASE_URL}/club/{CLUB_NAME}/members"
    print(f"[INFO] Fetching members from {CLUB_NAME}...")
    club_response = requests.get(club_url, headers=HEADERS); club_response.raise_for_status()
    members = sorted([member['username'] for category in club_response.json().values() for member in category])
    total_members_to_process = len(members)
    print(f"[INFO] Found {total_members_to_process} members.")
    if MEMBER_LIMIT:
        members = members[:MEMBER_LIMIT]; total_members_to_process = len(members)
        print(f"[INFO] Processing a limited list of {total_members_to_process} members.")
except requests.exceptions.RequestException as e:
    print(f"[ERROR] Failed to fetch club members: {e}"); members = []

member_stats = []
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'club_stats_{CLUB_NAME}_{timestamp}.xlsx'

if members:
    for idx, username in enumerate(members):
        print(f"[INFO] Processing {idx+1}/{total_members_to_process}: {username}")
        try:
            data = process_member(username)
            if data: member_stats.append(data)
            generate_excel_report(member_stats, output_file)

            elapsed_total = time.time() - start_time
            avg_time_per_member = elapsed_total / (idx + 1)
            est_remaining = avg_time_per_member * (total_members_to_process - idx - 1)
            print(f"[INFO] > Done. Report updated. Games: {data.get('total_games', 0)}. ETA: {est_remaining/60:.1f} mins.")
        except Exception as e:
            print(f"[WARNING] An unexpected error occurred for user {username}: {e}")

print(f"\n[SUCCESS] Analysis complete in {(time.time() - start_time)/60:.2f} minutes.")
print(f"[SUCCESS] Final report saved to '{output_file}'")
